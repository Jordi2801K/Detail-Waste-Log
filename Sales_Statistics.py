import pandas as pd
import streamlit as st
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font
import io

# # 데이터 프레임으로 읽는 함수
def read_df(file):
    # 엑셀 파일 불러오기 전, 필요한 탭만 추출하기 위해 탭의 리스트를 작성
    sheets = ["SUN", "MON", "TUE", "WED", "THUR", "FRI", "SAT"]

    # 업로드된 파일에서 필요한 탭만 불러오기
    df_dict = pd.read_excel(file, engine='pyxlsb', sheet_name=sheets)

    # 불러온 파일은 딕셔너리 타입이므로 각각의 변수로 매핑
    dfs = []
    for df in df_dict.values():
        dfs.append(df)
    
    return dfs


# # 데이터 프레임 처리 함수
def process_data(df, day_index):
    # 요일 리스트 (일요일부터 토요일까지)
    days = ['SUNDAY', 'MONDAY', 'TUESDAY', 'WEDNESDAY', 'THURSDAY', 'FRIDAY', 'SATURDAY']
    current_day = days[day_index]

    # Columns 숫자로 변경
    rename_col = range(len(df.columns))
    df.columns = rename_col

    # # 필요 없는 데이터들 제거
    # Tempreture 제거
    df = df[df.iloc[:, 0] != 'Tempreture']

    # NAME 제거
    df = df[df.iloc[:, 0] != 'NAME']

    # 표 이름 제거 (요일에 맞춰 자동으로 처리)
    df = df[df.iloc[:, 0] != f'ROLLER GRILL - {current_day}']
    df = df[df.iloc[:, 0] != f'BURRITOS - {current_day}']
    df = df[df.iloc[:, 0] != f'HOT TO GO - {current_day}']
    df = df[df.iloc[:, 0] != f'DELI EXPRESS - {current_day}']
    df = df[df.iloc[:, 0] != f'DELI EXPRESS / BIG AZ - {current_day}']
    df = df[df.iloc[:, 0] != '0x7']

    # NaN값 제거
    df = df.dropna(subset=[df.columns[0]])

    # 표의 아래 통계 계산해주는 셀 제거
    df = df[df.iloc[:, 0] != '0x7']
    df = df[df.iloc[:, 0] != 'ROLLER GRIL\nHOURS WASTE %']
    df = df[df.iloc[:, 0] != 'BURRITOS\nHOURS WASTE %']
    df = df[df.iloc[:, 0] != 'PAPA PRIMOS\nHOURS WASTE %']
    df = df[df.iloc[:, 0] != 'DELI EXPRESS\nHOURS WASTE %']
    df = df[df.iloc[:, 0] != f'TOTAL {current_day}\nHOURS WASTE %']
    df = df[df.iloc[:, 0] != 'TOTAL SUNDAY\nHOURS WASTE %']
    
    # 추가적인 총계 행 제거
    for day in days:
        df = df[df.iloc[:, 0] != f'TOTAL {day}']
        df = df[df.iloc[:, 0] != 'HOURS WASTE %']

    # ITEM # 제거
    df = df.drop(columns = [1, 50, 51, 52], axis=1)

    # Columns 변동되었으므로 한번 더 변경
    rename_col = range(len(df.columns))
    df.columns = rename_col

    # 셀에 값이 입력되면 생기는 값이 숫자인 행을 제거
    df = df[~df.iloc[:, 0].apply(lambda x: isinstance(x, (int, float)) or str(x).isdigit())]

    # 인덱스 재설정
    df.reset_index(drop=True, inplace=True)

    return df


# # TOTAL WEEK BY ITEM 계산 함수
def cal_total_week_by_item(dfs):
    items = {}
    days = ['Sunday', 'Monday', 'Tuesday', 'Wednesday', 'Thursday', 'Friday', 'Saturday']

    for day_index, df in enumerate(dfs):
        for i in range(len(df)):
            item_name = df.iloc[i, 0]
            if item_name not in items:
                items[item_name] = {day: {'time_disposal': [0]*24, 'time_put': [0]*24, 'time_waste': ['0%']*24, 'total_waste': '0%'} for day in days}

            for j in range(1, len(df.columns)):
                if pd.notna(df.iloc[i, j]):
                    hour = (j - 1) // 2
                    if (j % 2) == 1:  # 홀수 열 처리 (즉, put 열)
                        items[item_name][days[day_index]]['time_put'][hour] += float(df.iloc[i, j])
                    else:  # 짝수 열 처리 (즉, disposal 열)
                        items[item_name][days[day_index]]['time_disposal'][hour] += float(df.iloc[i, j])

    # Calculate waste percentages and totals
    for item in items:
        for day in days:
            day_data = items[item][day]
            for hour in range(24):
                if hour >= 4:
                    if day_data['time_put'][hour - 4] != 0:
                        waste = (day_data['time_disposal'][hour] / day_data['time_put'][hour - 4]) * 100
                        day_data['time_waste'][hour] = f"{waste:.1f}%"
                    else:
                        day_data['time_waste'][hour] = "0%"
                else:
                    day_data['time_waste'][hour] = "0%"

            total_disposal = sum(day_data['time_disposal'][4:])
            total_put = sum(day_data['time_put'][:20])
            if total_put != 0:
                total_waste = (total_disposal / total_put) * 100
                day_data['total_waste'] = f"{total_waste:.1f}%"
            else:
                day_data['total_waste'] = "0%"

    return items


# # TOTAL WASTE BY ITEM 엑셀 파일로 변환하는 함수
def create_excel_file_by_item(waste_by_item):
    wb = Workbook()
    ws = wb.active
    ws.title = "Waste Log Summary by Item"
    days = ['Sunday', 'Monday', 'Tuesday', 'Wednesday', 'Thursday', 'Friday', 'Saturday']
    
    current_row = 1

    for item, data in waste_by_item.items():
        # Write item name with larger font and bold
        ws.cell(row=current_row, column=1, value=f"Item: {item}")
        ws.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=35)
        cell = ws.cell(row=current_row, column=1)
        cell.font = Font(size=14, bold=True)
        cell.alignment = Alignment(horizontal="center", vertical="center")
        current_row += 2  # Add an extra empty row after the item name

        # Write day names
        for day_index, day in enumerate(days):
            start_col = 1 + day_index * 5
            ws.cell(row=current_row, column=start_col, value=day)
            ws.merge_cells(start_row=current_row, start_column=start_col, end_row=current_row, end_column=start_col + 3)
            ws.cell(row=current_row, column=start_col).alignment = Alignment(horizontal="center", vertical="center")
        current_row += 1

        # Write headers
        headers = ['Hour', 'Disposal', 'Put', 'Waste (%)']
        for day_index in range(7):
            for header_index, header in enumerate(headers):
                cell = ws.cell(row=current_row, column=1 + day_index * 5 + header_index, value=header)
                cell.alignment = Alignment(horizontal="center", vertical="center")
        current_row += 1

        # Write data
        for hour in range(24):
            for day_index, day in enumerate(days):
                for col_offset in range(4):
                    cell = ws.cell(row=current_row, column=1 + day_index * 5 + col_offset)
                    if col_offset == 0:
                        cell.value = hour + 1
                    elif col_offset == 1:
                        cell.value = data[day]['time_disposal'][hour]
                    elif col_offset == 2:
                        cell.value = data[day]['time_put'][hour]
                    else:
                        cell.value = data[day]['time_waste'][hour]
                    cell.alignment = Alignment(horizontal="center", vertical="center")
            current_row += 1

        # Write total waste
        for day_index, day in enumerate(days):
            cell_total = ws.cell(row=current_row, column=1 + day_index * 5, value='Total Waste (%)')
            cell_total.alignment = Alignment(horizontal="center", vertical="center")
            cell_value = ws.cell(row=current_row, column=4 + day_index * 5, value=data[day]['total_waste'])
            cell_value.alignment = Alignment(horizontal="center", vertical="center")

        # Add empty rows between items
        current_row += 4  # Increased the number of empty rows for better separation

    # Save to a BytesIO object
    excel_file = io.BytesIO()
    wb.save(excel_file)
    excel_file.seek(0)
    
    return excel_file




## # 메인 화면

st.title("A detailed summary of WASTE LOG")
st.subheader("For Troop Mini Mall", divider=True)

file = st.file_uploader("Upload excel file", type=["xlsb"])

# 파일이 업로드된 경우 데이터 프레임으로 읽기
if file :
    # 데이터 프레임 읽기
    dfs = read_df(file)

    # 각 데이터 프레임 처리
    for i in range(7):
        dfs[i] = process_data(dfs[i], i)  # i는 0부터 6까지의 요일 인덱스

    # TOTAL WEEK BY ITEM 계산
    waste_by_item = cal_total_week_by_item(dfs)

    # TOTAL WEEK BU ITEM 엑셀 파일로 변환
    excel_file_by_item = create_excel_file_by_item(waste_by_item)

    # 다운로드 버튼 생성
    st.download_button(
        label="Download daily&hourly TOTAL WEEK BY ITEM Excel file",
        data=excel_file_by_item,
        file_name="waste_log_summary_by_item.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )

    # 데이터 미리보기
    st.header("Preview")
    
    days = ['Sunday', 'Monday', 'Tuesday', 'Wednesday', 'Thursday', 'Friday', 'Saturday']
    tabs = st.tabs(days)
    
    for day_index, tab in enumerate(tabs):
        with tab:
            st.subheader(f"{days[day_index]} Data")
            
            # 처음 2개 아이템만 선택
            preview_items = list(waste_by_item.items())[:2]
            
            for item, data in preview_items:
                st.write(f"**Item: {item}**")
                
                df = pd.DataFrame({
                    'Hour': range(1, 25),
                    'Disposal': data[days[day_index]]['time_disposal'],
                    'Put': data[days[day_index]]['time_put'],
                    'Waste (%)': data[days[day_index]]['time_waste']
                })
                
                st.table(df)
                
                st.write(f"Total Waste: {data[days[day_index]]['total_waste']}")
                st.write("---")  # 구분선 추가
